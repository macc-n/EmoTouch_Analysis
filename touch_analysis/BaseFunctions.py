# -*- coding: utf-8 -*-
"""
Created on Wed May 29 17:02:49 2019

@author: robby
"""
import pandas as pd
import numpy as np
import math
import random 

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    import pandas as pd

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist 
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()

"""
custom train_test_split per splittare il dataset basandomi sui pattern e non sugli swipe (usati per la verifica)

"""
def train_test_split_custom(dataset, y, test_size, sampleNumber=0, leaveOneOut=False):
    if leaveOneOut == False:
        test_size = test_size*100
        sampleNumber = math.ceil(10*test_size/100)
        print(sampleNumber)
        indicies = random.sample(range(0, 10), sampleNumber)
    else:
        indicies = [sampleNumber]
        
    print("Indicies: "+str(indicies))
    train_ = dataset.copy()

    test_ = pd.DataFrame(columns = dataset.columns)
    y_train_ = pd.DataFrame(columns = ['Matricola'])
    y_test_ = pd.DataFrame(columns = ['Matricola'])
    
    for index in indicies:
        #test_ = test_.append(train_[train_['Task'].str.contains("_"+str(index)+"_")], sort=False)
        train_ = train_[train_['Task'].str.contains("_"+str(index)+"_")==False]
    
    test_ = dataset.loc[dataset.index.difference(train_.index)]

    print(test_['Task'])

    y_train_ = dataset.iloc[train_.index]
    y_train_ = y_train_[['Matricola']]
    y_test_ = dataset.iloc[test_.index]
    y_test_ = y_test_[['Matricola']]
    
    #print(test_['Matricola'].drop_duplicates(keep='first'))
    
    return train_, test_, y_train_, y_test_
"""
aggiunge primo e ultimo punto al dataframe df
input: dfActions, dataframe con tutti i tocchi
        df, dataframe interessato all'aggiunta
        add 'coords' aggiunge le coordinate
        add 'values' aggiunge i valori tetha
output: dataframe con i nuovi dati
"""
def addFirstLastPoints(dfActions, df, add):
    dfCopy = df.copy()
    if add == 'coords':
        temp = pd.DataFrame([[dfActions.iloc[0][2], 
                              dfActions.iloc[0][3], 
                              dfActions.iloc[0]['Time'], 
                              dfActions.iloc[0]['oldIndex']]], 
                            columns=df.columns)
        dfCopy = dfCopy.append(temp, ignore_index=True, sort=False)
        temp = pd.DataFrame([[dfActions.iloc[len(dfActions.index)-1][2], 
                              dfActions.iloc[len(dfActions.index)-1][3], 
                              dfActions.iloc[len(dfActions.index)-1]['Time'], 
                              dfActions.iloc[len(dfActions.index)-1]['oldIndex']]], 
                            columns=df.columns)
        dfCopy = dfCopy.append(temp, ignore_index=True, sort=False)
    elif add == 'values' and len(dfCopy.index)>1:
        temp = pd.DataFrame([[dfCopy.iloc[1][0],
                              dfActions.iloc[0]['Time']]], columns=df.columns)
        dfCopy = dfCopy.append(temp, ignore_index=True, sort=False)    
        temp = pd.DataFrame([[dfCopy.iloc[len(dfCopy.index)-3][0],
                              dfActions.iloc[len(dfActions.index)-1]['Time']]], columns=df.columns)
        dfCopy = dfCopy.append(temp, ignore_index=True, sort=False)
        
    dfCopy = dfCopy.sort_values(by=['Time'])
    dfCopy = dfCopy.reset_index(drop=True)
    return dfCopy

"""
calcola il displacemente dei tocchi nel dataframe in input
outPut: valore del displacement
"""
def calcolaDisplacemente(dfActions):
    displacement = 0
    
    crdXArray = dfActions.iloc[:,2].values
    crdYArray = dfActions.iloc[:,3].values
    
    startX = dfActions.iloc[0][2]
    startY = dfActions.iloc[0][3]
    
    for crdX, crdY in zip(crdXArray, crdYArray):
        tempDisplacement = math.sqrt(((crdX-startX)**2 + (crdY-startY)**2))
        displacement += tempDisplacement
        startX = crdX
        startY = crdY
    
    return displacement

"""
calcola la distanza tra 2 punti
"""
def getDistance(pointA, pointB):
    distance = math.sqrt(((pointB[0]-pointA[0])**2 + (pointB[1]-pointA[1])**2))
    return distance

"""
calcola i vaolori da utilizzare per la percentuale di una features,
input: dfActions, dataframe dei tocchi
        intervalArray indica le percentuali di riferimento (es. 20%, 50%, 80%)
output: indici dei valori che fanno riferimento all'x% dell'intervallo
"""
def calcolaIntervalli(dfActions, intervalArray):
    numActions = len(dfActions.index)        
    #intervalArray = [20, 50, 80, 100]
    percentageProgress = []
    for value in intervalArray:
        temp = round((numActions/100)*value)
        if(value == intervalArray[0] and temp == 0):
            temp = temp + 1
        percentageProgress.append(temp)
    return percentageProgress
"""
calcola la direzione tra due punti
input: intevallo: dataframe dei tocchi
output: dataframe contenti i tetha per le coppie di tocchi
"""
#mode: sp per il calcolo start point - point, pp per il calcolo point-point, ep per end point - point
def angleDirectionTwoPoints(intervallo, mode):
    idx = 0
    oldIdx = 0
    df = intervallo.copy() #2-3-Time-oldIndex
    df[2] = df[2].astype(np.float64)
    df[3] = df[3].astype(np.float64)
    tethaDF = pd.DataFrame(columns = ['tetha','direction'])
    #print("df")
    #print(df)
    while idx < len(df.index)-1:
        if mode == 'sp':
            idx = 0            
        cx1 = df.iloc[idx][0]
        cy1 = df.iloc[idx][1]
        if mode == 'sp':
            idx = oldIdx
        elif mode == 'ep':
            oldIdx = idx
            idx = len(df.index)-2
        cx2 = df.iloc[idx+1][0]
        cy2 = df.iloc[idx+1][1]
        if mode == 'ep':
            idx = oldIdx
        
        if (cx2-cx1 == 0):
            #print("ZERO")
            angolo = np.NaN
        else:
            angolo = math.degrees(np.arctan((cy2-cy1)/(cx2-cx1)))
        
        if cx2 < cx1:
            angolo += 180
            
        if angolo < 0:
            angolo += 360
            
        direction = getDirection(angolo)
        
        temp = pd.DataFrame([[angolo,direction]], columns = tethaDF.columns)
        tethaDF = tethaDF.append(temp, ignore_index=True, sort=False)
        idx += 1
        oldIdx = idx
    return tethaDF

"""
indica la direzione (da 1 a 8) del vettore
"""
def getDirection(angolo):
    inf = 0
    sup = 45
    dir = 1
    trovato = False
    while dir <= 8 and trovato == False:
        if inf <= angolo < sup:
            trovato = True
        else:
            inf = sup
            sup += 45
            dir += 1
    return dir

"""
restituisce un subDataFrame contenente i valori tra i 2 limini del dataframe originario
"""
def getSubActionsInfo(actionsInfo, oldIndexInf, oldIndexSup):
    subActionsInfo = actionsInfo[(actionsInfo.oldIndex >= oldIndexInf) & (actionsInfo.oldIndex <= oldIndexSup)]
    return subActionsInfo

"""
calcola l'angolo tra tre punti
input: intervallo, dataframe dei tocchi
        filter, indica l'angolo massimo dei tetha ricercato (default 360 -> nessun filtro)
        asse, se si vuole calcolare il tetha rispetto asse x o y

"""
def anglesBetweenThreePoints(intervallo, filter=360, asse=''):
    idx = 0
    oldIdx = 0
    #print("intervallo")
    #print(intervallo)
    df = intervallo.copy()
    if asse == 'x':
        df = df.append(pd.DataFrame([[0,0,-1,-1]], columns = df.columns), sort=False)
        df = df.append(pd.DataFrame([[100,0,-1,-1]], columns = df.columns), sort=False)
        df = df.sort_values(by=['Time'])
        df = df.reset_index(drop=True)
    elif asse == 'y':
        df = df.append(pd.DataFrame([[0,0,-1,-1]], columns = df.columns), sort=False)
        df = df.append(pd.DataFrame([[0,100,-1,-1]], columns = df.columns), sort=False)
        df = df.sort_values(by=['Time'])
        df = df.reset_index(drop=True)

    if filter == 360:
        tethaDF = pd.DataFrame(columns = ['tetha','Time'])
    else:
        tethaDF = pd.DataFrame(columns = ['X','Y','Time','oldIndex'])
    
    while (idx < len(df.index) - 2):
        if asse == 'x' or asse == 'y':
            idx = 0
#        print(df)
        cx1 = df.iloc[idx][0]
        cx2 = df.iloc[idx+1][0]
        cy1 = df.iloc[idx][1]
        cy2 = df.iloc[idx+1][1]
        
        if asse == 'x' or asse == 'y':
            idx = oldIdx
            
        cx3 = df.iloc[idx+2][0]
        cy3 = df.iloc[idx+2][1]

        vec1, vec2 = readCoords(np.array([cx1,cy1]), np.array([cx2,cy2]), np.array([cx3,cy3]))
        ris = 180 - tetha(vec1, vec2)
        
        if filter != 360:       
            if (ris <= filter):
                temp = pd.DataFrame([[cx2, cy2, df.iloc[idx+1]['Time'], df.iloc[idx+1]['oldIndex']]], columns = ['X','Y','Time','oldIndex'])
                tethaDF = tethaDF.append(temp, ignore_index=True, sort=False)   
                idx = idx + 1 
            else:
                df = df[(df.X != cx2) | (df.Y != cy2)]
        else:
            if asse == 'x' or asse == 'y':
                indice = idx+2
            else:
                indice = idx+1
            temp = pd.DataFrame([[ris, df.iloc[indice]['Time']]], columns=['tetha','Time'])
            tethaDF = tethaDF.append(temp,ignore_index=True, sort=False)
            idx = idx + 1
            oldIdx = idx

    #input("wait..")
    return tethaDF   

"""
legge le coordinate delle 3 rows in input
"""
def readCoords(val1, val2, val3):
    
    x1 = val2[0] - val1[0]
    y1 = val2[1] - val1[1]
    vec1 = np.array([x1, y1])
    
    x2 = val3[0] - val2[0]
    y2 = val3[1] - val2[1]
    vec2 = np.array([x2,y2])
    return vec1,vec2

"""
calcola il tetha tra i 3 punti
input: vec1, vettore contente i primi 2 punti 
    vec2, vettore contenente il 2° punto e il 3°
"""
def tetha(vec1,vec2):
    
    prodScal =  vec1[0]*vec2[0] + vec1[1]*vec2[1] 
    normVec1 =  np.linalg.norm(vec1)
    normVec2 =  np.linalg.norm(vec2)
    
    t = prodScal / (normVec1*normVec2)
    degrees = math.degrees(np.arccos(float(t)))
    return degrees 

"""
permette l'aggiunte del deltaTime tra i tocchi per una migliore visione dei punti nel tempo
viene anche resetto l'index, mantendo però il vecchio index nella colonna oldIndex (per ricollegare facilmente
gli anglePoints al dataframe originario dei tocchi)
"""

def aggiungiDeltaTimeTraPunti(dfActionsOld):
    
    dfActions = dfActionsOld.copy()
    startT = dfActions.iloc[0][1]    
    deltaTSerie = pd.Series([])
    deltaTSerie = deltaTSerie.append(pd.Series([0]), ignore_index = True)
    
    for t in zip(dfActions.iloc[1:,1]):
        deltaTSerie = deltaTSerie.append(pd.Series(t-startT), ignore_index = True)
    
    #deltaTSerie.index += 506
    dfActions = dfActions.assign(**{'Time':deltaTSerie.values})
    
    dfActions = dfActions.reset_index()
    dfActions = dfActions.assign(**{'oldIndex':dfActions.iloc[:,0]})
    dfActions = dfActions.drop(['index'], axis=1)
    
    return dfActions
    